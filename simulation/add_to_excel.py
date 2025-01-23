import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .get_probability import get_avarge_waiting_time_for_custumer,get_expected_service_time,get_probability_server_idel,get_probability_that_customer_waits

###### add to excel
def save_data_to_excel(file_name, time_between_arrivals_data, service_time_data, simulation_table, service2_time_data=None):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Simulation Data"

    # عنوان الصفحة
    sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=12)
    title_cell = sheet.cell(row=1, column=2, value="Simulation Tables in Excel")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # لون أزرق فاتح

    start_col_table1 = 2
    start_col_table2 = 7

    # إضافة الجدول الأول والثاني
    add_table_with_title(sheet, "arrival probability", time_between_arrivals_data, start_row=4, start_col=start_col_table1)
    add_table_with_title(sheet, "Service probability", service_time_data, start_row=4, start_col=start_col_table2)

    # إضافة الجدول الثالث (Service 2) إذا كان يحتوي على بيانات
    start_row_table3 = 4  # البداية الافتراضية
    if service2_time_data and any(service2_time_data.values()):
        start_col_table3 = start_col_table2 + len(service_time_data.keys()) + 2  # العمود بجانب الجدول الثاني
        add_table_with_title(sheet, "Service 2 probability", service2_time_data, start_row=4, start_col=start_col_table3)
        # تحديث آخر صف مستخدم للجدول الثالث
        start_row_table3 = max(get_last_row(sheet, start_col_table3), start_row_table3)

    # تحديد موقع الجدول الرابع (simulation table) بعد آخر صف في الجداول الثلاثة + صفين
    last_row_table1 = get_last_row(sheet, start_col_table1)
    last_row_table2 = get_last_row(sheet, start_col_table2)
    last_row_table3 = get_last_row(sheet, start_col_table3 if service2_time_data else start_col_table2)

    start_row_simulation_table = max(last_row_table1, last_row_table2, last_row_table3) + 3
    add_table_with_title(sheet, "simulation table", simulation_table, start_row=start_row_simulation_table, start_col=2)

    # ضبط عرض الأعمدة
    auto_adjust_column_width(sheet)

    # إضافة القيم النهائية
    add_summary_values(sheet, simulation_table, service_time_data)

    # حفظ الملف
    wb.save(file_name)



def add_summary_values(sheet, simulation_table, server_01):
    """إضافة القيم النهائية إلى آخر صف في الملف."""
    last_row = get_last_row(sheet, 2) + 1  # الحصول على آخر صف في العمود الثاني
    summary_data = {
        "Average Waiting Time for Customers": get_avarge_waiting_time_for_custumer(simulation_table),
        "Probability That a Customer Waits": get_probability_that_customer_waits(simulation_table),
        "Probability Server Idle": get_probability_server_idel(simulation_table),
        "Expected Service Time": get_expected_service_time(server_01),
    }

    # إضافة القيم
    for row_offset, (label, value) in enumerate(summary_data.items()):
        label_cell = sheet.cell(row=last_row + row_offset, column=2, value=label)
        value_cell = sheet.cell(row=last_row + row_offset, column=3, value=value)

        # تنسيق الخلايا
        for cell in (label_cell, value_cell):
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # لون أزرق فاتح
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
def get_last_row(sheet, start_col):
    """الحصول على آخر صف يحتوي على بيانات في عمود معين."""
    last_row = 1
    for row in sheet.iter_rows(min_col=start_col, max_col=start_col):
        for cell in row:
            if cell.value:
                last_row = max(last_row, cell.row)
    return last_row
def auto_adjust_column_width(sheet):
    """ضبط عرض الأعمدة بناءً على أكبر كلمة في العمود."""
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # الحصول على حرف العمود
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # إضافة مسافة صغيرة
        sheet.column_dimensions[col_letter].width = adjusted_width
def add_table_with_title(sheet, title, data, start_row, start_col):
    # إضافة العنوان
    title_cell = sheet.cell(row=start_row, column=start_col, value=title)
    title_cell.font = Font(bold=True, size=15)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # لون رمادي فاتح

    # إضافة الرؤوس
    headers = list(data.keys())
    for col_num, header in enumerate(headers, start=start_col):
        cell = sheet.cell(row=start_row + 1, column=col_num, value=header)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # لون خلفية أصفر
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # إضافة البيانات
    for row_num, row_data in enumerate(zip(*data.values()), start=start_row + 2):
        for col_num, value in enumerate(row_data, start=start_col):
            if isinstance(value, tuple):
                value = str(value)
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # لون رمادي فاتح
